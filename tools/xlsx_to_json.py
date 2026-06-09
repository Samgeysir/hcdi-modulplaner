"""
xlsx_to_json.py — wandelt die ausgefüllte Excel-Vorlage in eine
data/extra_modules_<Semester>.json um (Schema, das der Modulplaner erwartet).

Aufruf (Dev, braucht openpyxl):
    python3 tools/xlsx_to_json.py <ausgefuelltes.xlsx> <Semester> [Zielordner]

Beispiel:
    python3 tools/xlsx_to_json.py Vorlage_Zusatzmodule.xlsx 26HS
    -> schreibt data/extra_modules_26HS.json

Liest das Tabellenblatt »Module«. Beispiel-/Leerzeilen werden übersprungen.
Spalten werden über die deutsche Überschrift (Zeile 1) den JSON-Keys zugeordnet,
die Reihenfolge der Spalten ist daher egal.
"""

import os
import sys
import json

from openpyxl import load_workbook

# Deutsche Überschrift -> JSON-Key (muss mit build_template.py übereinstimmen).
HEADER_TO_KEY = {
    "Hochschule / Anbieter": "Hochschule",
    "Modultitel (Deutsch)": "title",
    "ECTS": "ects",
    "Studiengang(e)": "studyPrograms",
    "Modultitel (Englisch)": "titleEN",
    "Unterrichtssprache": "language",
    "Dozierende": "teachers",
    "Modulverantwortung": "moduleResponsibles",
    "Standort / Ort": "locations",
    "Modultyp": "moduleTypes",
    "Modulinhalt / Beschreibung": "courseContent",
    "Kernidee": "keyIdea",
    "Kompetenzen": "competences",
    "Literatur": "literature",
    "Voraussetzungen": "requirements",
    "Leistungsnachweis": "assessment",
    "Link (URL)": "url",
    "Wochentag": "lektionDayOfWeek",
    "Zeit von": "lektionTimeFrom",
    "Zeit bis": "lektionTimeTo",
    "Erster Termin": "lektionFirstDate",
    "Letzter Termin": "lektionLastDate",
    "Alle Termine": "lektionDates",
    "Raum / Räume": "lektionRooms",
}

WEEKDAY_DE_EN = {
    "montag": "Monday", "dienstag": "Tuesday", "mittwoch": "Wednesday",
    "donnerstag": "Thursday", "freitag": "Friday", "samstag": "Saturday",
    "sonntag": "Sunday",
}

REQUIRED = ["Hochschule", "title", "ects", "studyPrograms"]


def _clean_header(h):
    return (str(h).replace("*", "").strip()) if h is not None else ""


def _as_text(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _norm_date(v):
    """Excel-Datum/Text -> TT.MM.JJJJ."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%d.%m.%Y")
    return str(v).strip()


def convert(xlsx_path, semester, out_dir):
    wb = load_workbook(xlsx_path, data_only=True)
    if "Module" not in wb.sheetnames:
        sys.exit("Fehler: Tabellenblatt »Module« nicht gefunden.")
    ws = wb["Module"]

    # Kopfzeile -> Spaltenindex je JSON-Key
    headers = {}
    unknown = []
    for col, cell in enumerate(ws[1], start=1):
        name = _clean_header(cell.value)
        if not name:
            continue
        key = HEADER_TO_KEY.get(name)
        if key:
            headers[key] = col
        else:
            unknown.append(name)
    missing = [HEADER_TO_KEY[h] for h in HEADER_TO_KEY if HEADER_TO_KEY[h] in REQUIRED and HEADER_TO_KEY[h] not in headers.values()]
    missing_req = [k for k in REQUIRED if k not in headers]
    if missing_req:
        sys.exit(f"Fehler: Pflichtspalten fehlen in der Vorlage: {missing_req}")
    if unknown:
        print(f"Hinweis: unbekannte Spalten ignoriert: {unknown}")

    records = []
    skipped = 0
    for row in ws.iter_rows(min_row=2):
        def val(key):
            col = headers.get(key)
            return row[col - 1].value if col else None

        title = _as_text(val("title"))
        hochschule = _as_text(val("Hochschule"))
        # Beispiel-/Leerzeilen überspringen
        if not title or not hochschule:
            skipped += 1
            continue
        low = title.lower()
        if (low.startswith("beispiel")
                or low.startswith("testmodul a (extern")
                or low.startswith("testmodul b (extern")):
            # Vorlagen-Beispielzeilen nicht übernehmen
            skipped += 1
            continue

        rec = {}
        for key, col in headers.items():
            raw = row[col - 1].value
            if key == "ects":
                try:
                    rec["ects"] = int(float(raw)) if raw not in (None, "") else ""
                except (TypeError, ValueError):
                    rec["ects"] = _as_text(raw)
            elif key == "lektionDayOfWeek":
                rec[key] = WEEKDAY_DE_EN.get(_as_text(raw).lower(), _as_text(raw))
            elif key in ("lektionFirstDate", "lektionLastDate"):
                rec[key] = _norm_date(raw)
            elif key == "lektionDates":
                txt = _as_text(raw)
                parts = [p.strip() for p in txt.replace(";", ",").split(",") if p.strip()]
                rec[key] = ", ".join(parts)
            else:
                rec[key] = _as_text(raw)

        # Dozierende auch als detailedLecturers (für die Kartenanzeige)
        if rec.get("teachers") and not rec.get("detailedLecturers"):
            rec["detailedLecturers"] = rec["teachers"]
        records.append(rec)

    safe = "".join(c for c in semester if c.isalnum())
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"extra_modules_{safe}.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    print(f"{len(records)} Module geschrieben -> {out_path}  ({skipped} Zeilen übersprungen)")


def main():
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    xlsx_path = sys.argv[1]
    semester = sys.argv[2]
    repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_dir = sys.argv[3] if len(sys.argv) > 3 else os.path.join(repo_root, "data")
    convert(xlsx_path, semester, out_dir)


if __name__ == "__main__":
    main()
