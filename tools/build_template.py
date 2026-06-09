"""
build_template.py — erzeugt die Excel-Vorlage `Vorlage_Zusatzmodule.xlsx`.

Die Vorlage geben die Verantwortlichen ausgefüllt zurück; danach wandelt
`xlsx_to_json.py` sie in `data/extra_modules_<sem>.json` um.

Aufruf (Dev, braucht openpyxl):  python3 tools/build_template.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

# (Deutsche Spaltenüberschrift, JSON-Key, Pflicht?, Spaltenbreite, Beispiel A, Beispiel B, Format/Hinweis)
COLUMNS = [
    ("Hochschule / Anbieter", "Hochschule", True, 28, "Externer Studiengang", "Externer Studiengang", "Name der Hochschule bzw. des anbietenden Studiengangs."),
    ("Modultitel (Deutsch)", "title", True, 32, "Testmodul A (extern)", "Testmodul B (extern, ohne feste Zeit)", "Voller Modultitel auf Deutsch."),
    ("ECTS", "ects", True, 8, 3, 6, "Anzahl ECTS-Kreditpunkte (Zahl)."),
    ("Studiengang(e)", "studyPrograms", True, 28, "Externer Beispiel-Studiengang", "Externer Beispiel-Studiengang", "Studiengang(e), zu denen das Modul gehört. Mehrere mit Komma trennen."),
    ("Modultitel (Englisch)", "titleEN", False, 32, "Test Module A (external)", "Test Module B (external)", "Optional. Englischer Titel."),
    ("Unterrichtssprache", "language", False, 16, "Deutsch", "Deutsch", "Deutsch / Englisch / Bilingual."),
    ("Dozierende", "teachers", False, 24, "Vorname Nachname", "Andere Person", "Name(n) der Dozierenden. Mehrere mit Komma trennen."),
    ("Modulverantwortung", "moduleResponsibles", False, 24, "Vorname Nachname", "Andere Person", "Optional. Modulverantwortliche Person."),
    ("Standort / Ort", "locations", False, 16, "Olten", "Brugg-Windisch", "Optional. Durchführungsort."),
    ("Modultyp", "moduleTypes", False, 16, "Wahlmodul", "Pflichtmodul", "Optional, z.B. Pflicht- / Wahlmodul."),
    ("Modulinhalt / Beschreibung", "courseContent", False, 40, "Kurzbeschreibung des Modulinhalts …", "Kurzbeschreibung des Modulinhalts …", "Fliesstext. Beschreibung des Inhalts."),
    ("Kernidee", "keyIdea", False, 30, "", "", "Optional. Kernidee / Leitgedanke."),
    ("Kompetenzen", "competences", False, 30, "Beispielkompetenz 1, Beispielkompetenz 2", "", "Optional. Erworbene Kompetenzen."),
    ("Literatur", "literature", False, 24, "", "", "Optional. Literaturangaben."),
    ("Voraussetzungen", "requirements", False, 24, "Keine besonderen Voraussetzungen.", "", "Optional. Voraussetzungen."),
    ("Leistungsnachweis", "assessment", False, 22, "Schriftliche Prüfung", "Projektarbeit", "Optional. Art des Leistungsnachweises."),
    ("Link (URL)", "url", False, 28, "https://www.fhnw.ch", "https://www.fhnw.ch", "Optional. Link zur Modulbeschreibung."),
    # --- Stundenplan (optional; nur ausfüllen, wenn das Modul im Kalender erscheinen soll) ---
    ("Wochentag", "lektionDayOfWeek", False, 14, "Montag", "", "Optional (Stundenplan). Wochentag aus der Auswahlliste."),
    ("Zeit von", "lektionTimeFrom", False, 10, "10:15", "", "Optional (Stundenplan). Format HH:MM, z.B. 10:15."),
    ("Zeit bis", "lektionTimeTo", False, 10, "12:00", "", "Optional (Stundenplan). Format HH:MM, z.B. 12:00."),
    ("Erster Termin", "lektionFirstDate", False, 14, "14.09.2026", "", "Optional (Stundenplan). Format TT.MM.JJJJ."),
    ("Letzter Termin", "lektionLastDate", False, 14, "21.12.2026", "", "Optional (Stundenplan). Format TT.MM.JJJJ."),
    ("Alle Termine", "lektionDates", False, 40, "14.09.2026; 21.09.2026; 28.09.2026", "", "Optional (Stundenplan). Alle Daten TT.MM.JJJJ, mit Semikolon (;) getrennt."),
    ("Raum / Räume", "lektionRooms", False, 18, "Beispielraum 1.234", "", "Optional (Stundenplan). Raum(e). Mehrere mit Komma trennen."),
]

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F2A44")      # dunkelblau
HDR_REQ_FILL = PatternFill("solid", fgColor="7A1F3D")  # weinrot für Pflichtfelder
SECTION_FILL = PatternFill("solid", fgColor="E8ECF5")
EXAMPLE_FILL = PatternFill("solid", fgColor="FFF6E5")
THIN = Side(style="thin", color="C8CEDA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build():
    wb = Workbook()

    # ---------------- Sheet 1: Anleitung ----------------
    ws = wb.active
    ws.title = "Anleitung"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 34

    ws["A1"] = "Zusatzmodule – Erfassungsvorlage"
    ws["A1"].font = Font(name=FONT, size=16, bold=True, color="1F2A44")
    ws["A2"] = ("Bitte im Tabellenblatt »Module« pro Modul eine Zeile ausfüllen. "
                "Die beiden Beispielzeilen vor dem Ausfüllen löschen.")
    ws["A2"].font = Font(name=FONT, size=11, italic=True)
    ws.merge_cells("A2:D2")
    ws["A4"] = ("Pflichtfelder sind rot markiert. Stundenplan-Felder nur ausfüllen, wenn das Modul "
                "mit Tag/Zeit im Kalender erscheinen soll – sonst leer lassen (erscheint dann als »Unregelmässig«).")
    ws["A4"].font = Font(name=FONT, size=10)
    ws.merge_cells("A4:D4")

    hdr = ["Spalte", "Pflicht", "Bedeutung", "Format / Beispiel"]
    ws.append([])  # row 5 leer
    r = 6
    for i, h in enumerate(hdr):
        c = ws.cell(row=r, column=i + 1, value=h)
        c.font = Font(name=FONT, bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER
    r += 1
    for label, key, req, _w, exA, _exB, note in COLUMNS:
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT, bold=req)
        ws.cell(row=r, column=2, value="Ja" if req else "optional").font = Font(name=FONT, color="7A1F3D" if req else "555555")
        ws.cell(row=r, column=3, value=note).font = Font(name=FONT)
        ws.cell(row=r, column=4, value=f"z.B. {exA}" if exA != "" else "").font = Font(name=FONT, color="555555")
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

    ws.append([])
    r += 1
    ws.cell(row=r, column=1, value="Zurückgeben an:").font = Font(name=FONT, bold=True)
    ws.cell(row=r, column=3, value="die für den Modulplaner verantwortliche Person (ausgefülltes Excel).").font = Font(name=FONT)

    # ---------------- Sheet 2: Module ----------------
    ms = wb.create_sheet("Module")
    ms.sheet_view.showGridLines = False

    # Kopfzeile
    for col, (label, key, req, width, exA, exB, note) in enumerate(COLUMNS, start=1):
        cell = ms.cell(row=1, column=col, value=label + (" *" if req else ""))
        cell.font = Font(name=FONT, bold=True, color="FFFFFF")
        cell.fill = HDR_REQ_FILL if req else HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        cell.comment = Comment(note, "Vorlage")
        ms.column_dimensions[cell.column_letter].width = width

    # Beispielzeilen (Zeile 2 + 3), klar als Beispiel markiert
    for ridx, exidx in ((2, 4), (3, 5)):
        for col, spec in enumerate(COLUMNS, start=1):
            val = spec[exidx]
            cell = ms.cell(row=ridx, column=col, value=val if val != "" else None)
            cell.font = Font(name=FONT, italic=True, color="9A6A00")
            cell.fill = EXAMPLE_FILL
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
    ms.cell(row=2, column=1).comment = Comment("BEISPIEL – vor dem Ausfüllen löschen.", "Vorlage")

    # Leere Eingabezeilen vorformatieren
    for ridx in range(4, 60):
        for col in range(1, len(COLUMNS) + 1):
            cell = ms.cell(row=ridx, column=col)
            cell.font = Font(name=FONT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER

    ms.freeze_panes = "A2"
    ms.row_dimensions[1].height = 42

    # Dropdowns (Data Validation)
    key_to_col = {spec[1]: idx for idx, spec in enumerate(COLUMNS, start=1)}

    def col_letter(key):
        return ms.cell(row=1, column=key_to_col[key]).column_letter

    dv_day = DataValidation(type="list",
                            formula1='"Montag,Dienstag,Mittwoch,Donnerstag,Freitag,Samstag,Sonntag"',
                            allow_blank=True)
    dv_day.error = "Bitte einen Wochentag aus der Liste wählen."
    dv_day.prompt = "Wochentag wählen (für den Stundenplan)."
    ms.add_data_validation(dv_day)
    dl = col_letter("lektionDayOfWeek")
    dv_day.add(f"{dl}2:{dl}500")

    dv_lang = DataValidation(type="list", formula1='"Deutsch,Englisch,Bilingual"', allow_blank=True)
    ms.add_data_validation(dv_lang)
    ll = col_letter("language")
    dv_lang.add(f"{ll}2:{ll}500")

    dv_ects = DataValidation(type="decimal", operator="between", formula1=0, formula2=60, allow_blank=True)
    dv_ects.error = "ECTS als Zahl (0–60)."
    ms.add_data_validation(dv_ects)
    el = col_letter("ects")
    dv_ects.add(f"{el}2:{el}500")

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Vorlage_Zusatzmodule.xlsx")
    wb.save(out)
    print("geschrieben:", out)


if __name__ == "__main__":
    build()
